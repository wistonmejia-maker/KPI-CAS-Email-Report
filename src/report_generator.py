"""
Report Generator - Generación de Reportes Excel
=================================================
Módulo para generar reportes Excel con múltiples hojas y formato profesional.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .config import (
    COLUMNS, WEEKLY_REPORTS_DIR, MONTHLY_REPORTS_DIR,
    get_current_date_str, get_current_week_str
)
from .metrics import MetricsCalculator
from .change_detector import ComparisonResult

logger = logging.getLogger(__name__)


class ExcelReportStyles:
    """Estilos para reportes Excel"""
    
    HEADER_FILL = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    
    ALERT_FILL = PatternFill(start_color="ea4335", end_color="ea4335", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="fbbc04", end_color="fbbc04", fill_type="solid")
    SUCCESS_FILL = PatternFill(start_color="34a853", end_color="34a853", fill_type="solid")
    
    TITLE_FONT = Font(bold=True, size=14, color="202124")
    SUBTITLE_FONT = Font(bold=True, size=12, color="5f6368")
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


class ExcelReportGenerator:
    """Generador de reportes Excel"""
    
    def __init__(self, df: pd.DataFrame, comparison: Optional[ComparisonResult] = None):
        """
        Inicializa el generador de reportes
        
        Args:
            df: DataFrame con oportunidades actuales
            comparison: Resultado de comparación (opcional)
        """
        self.df = df.copy()
        self.comparison = comparison
        self.metrics = MetricsCalculator(df)
        
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl no disponible. Usando formato básico.")
    
    def generate_weekly_report(self, output_path: Optional[Path] = None) -> Path:
        """
        Genera reporte semanal completo
        
        Args:
            output_path: Ruta de salida (opcional)
            
        Returns:
            Path al archivo generado
        """
        if output_path is None:
            filename = f"{get_current_week_str()}_weekly_report.xlsx"
            output_path = WEEKLY_REPORTS_DIR / filename
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Hoja 1: Resumen Ejecutivo
            self._write_summary_sheet(writer)
            
            # Hoja 2: Por Responsable
            self._write_responsible_sheet(writer)
            
            # Hoja 3: Por País
            self._write_market_sheet(writer)
            
            # Hoja 4: Por KPI
            self._write_kpi_sheet(writer)
            
            # Hoja 5: Por Stage
            self._write_stage_sheet(writer)
            
            # Hoja 6: Oportunidades a Actualizar
            self._write_attention_sheet(writer)
            
            # Hoja 7: Cambios (si hay comparación)
            if self.comparison:
                self._write_changes_sheet(writer)
            
            # Hoja 8: Datos Completos
            self._write_data_sheet(writer)
        
        logger.info(f"Reporte generado: {output_path}")
        return output_path
    
    def generate_responsible_reports(self, output_dir: Optional[Path] = None) -> List[Path]:
        """
        Genera reportes individuales por responsable
        
        Args:
            output_dir: Directorio de salida
            
        Returns:
            Lista de paths a los archivos generados
        """
        if output_dir is None:
            output_dir = WEEKLY_REPORTS_DIR / get_current_week_str() / "por_responsable"
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        generated_files = []
        
        for responsible in self.df[COLUMNS['responsible']].unique():
            if pd.isna(responsible) or responsible == 'Sin Asignar':
                continue
            
            # Filtrar datos del responsable
            resp_df = self.df[self.df[COLUMNS['responsible']] == responsible].copy()
            
            # Generar archivo
            safe_name = "".join(c for c in responsible if c.isalnum() or c in (' ', '-', '_')).rstrip()
            filename = f"{safe_name}_oportunidades.xlsx"
            filepath = output_dir / filename
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Resumen
                self._write_responsible_summary(writer, responsible, resp_df)
                
                # Oportunidades
                resp_df.to_excel(writer, sheet_name='Oportunidades', index=False)
                
                # Oportunidades a actualizar
                calc = MetricsCalculator(resp_df)
                attention = calc.get_opportunities_to_update()
                if len(attention) > 0:
                    attention.to_excel(writer, sheet_name='Por_Actualizar', index=False)
            
            generated_files.append(filepath)
            logger.info(f"  Generado: {filename}")
        
        return generated_files
    
    def _write_summary_sheet(self, writer: pd.ExcelWriter):
        """Escribe hoja de resumen ejecutivo"""
        summary = self.metrics.get_summary()
        
        data = [
            ['RESUMEN EJECUTIVO', ''],
            ['Fecha de Reporte', datetime.now().strftime('%Y-%m-%d %H:%M')],
            ['', ''],
            ['MÉTRICAS GENERALES', ''],
            ['Total Oportunidades', summary['total_opportunities']],
            ['Valor Total USD', f"${summary['total_usd']:,.2f}"],
            ['Promedio USD', f"${summary['avg_usd']:,.2f}"],
            ['', ''],
            ['ALERTAS', ''],
            ['Oportunidades Estancadas (>30 días)', summary['stagnant_count']],
            ['Oportunidades En Riesgo (<7 días)', summary['at_risk_count']],
            ['', ''],
            ['DISTRIBUCIÓN', ''],
            ['Responsables Únicos', summary['unique_responsibles']],
            ['Países/Mercados', summary['unique_markets']],
            ['Clientes Únicos', summary['unique_customers']],
        ]
        
        # Si hay comparación, agregar cambios
        if self.comparison:
            data.extend([
                ['', ''],
                ['CAMBIOS vs PERÍODO ANTERIOR', ''],
                ['Oportunidades Nuevas', self.comparison.summary['new_count']],
                ['Oportunidades Eliminadas', self.comparison.summary['removed_count']],
                ['Oportunidades Modificadas', self.comparison.summary['changed_count']],
                ['Cambio en USD', f"${self.comparison.summary['usd_change']:,.2f}"],
            ])
        
        df_summary = pd.DataFrame(data, columns=['Métrica', 'Valor'])
        df_summary.to_excel(writer, sheet_name='Resumen', index=False)
    
    def _write_responsible_sheet(self, writer: pd.ExcelWriter):
        """Escribe hoja de métricas por responsable"""
        df_resp = self.metrics.get_responsible_summary_df()
        df_resp.to_excel(writer, sheet_name='Por_Responsable', index=False)
    
    def _write_market_sheet(self, writer: pd.ExcelWriter):
        """Escribe hoja de métricas por mercado"""
        df_market = self.metrics.get_market_summary_df()
        df_market.to_excel(writer, sheet_name='Por_País', index=False)
    
    def _write_kpi_sheet(self, writer: pd.ExcelWriter):
        """Escribe hoja de métricas por KPI"""
        kpi_data = []
        kpi_metrics = self.metrics.get_kpi_metrics()
        
        for kpi, data in kpi_metrics.items():
            kpi_data.append({
                'KPI': kpi,
                'Total_Oportunidades': data['count'],
                'Total_USD': data['total_usd'],
                'Promedio_USD': data['avg_usd'],
                'Tipo': data['category_info'].get('type', 'N/A')
            })
        
        df_kpi = pd.DataFrame(kpi_data)
        df_kpi = df_kpi.sort_values('Total_Oportunidades', ascending=False)
        df_kpi.to_excel(writer, sheet_name='Por_KPI', index=False)
    
    def _write_stage_sheet(self, writer: pd.ExcelWriter):
        """Escribe hoja de métricas por stage"""
        df_stage = self.metrics.get_stage_distribution()
        df_stage.to_excel(writer, sheet_name='Por_Stage', index=False)
    
    def _write_attention_sheet(self, writer: pd.ExcelWriter):
        """Escribe hoja de oportunidades que requieren atención"""
        attention = self.metrics.get_opportunities_to_update()
        
        if len(attention) > 0:
            # Seleccionar columnas relevantes
            cols = [COLUMNS['id'], COLUMNS['responsible'], COLUMNS['market'],
                    COLUMNS['customer'], COLUMNS['stage'], COLUMNS['usd'],
                    COLUMNS['close_date'], 'Razon_Alerta', '_days_to_close']
            cols = [c for c in cols if c in attention.columns]
            
            attention[cols].to_excel(writer, sheet_name='Por_Actualizar', index=False)
        else:
            pd.DataFrame({'Mensaje': ['No hay oportunidades que requieran atención inmediata']}).to_excel(
                writer, sheet_name='Por_Actualizar', index=False
            )
    
    def _write_changes_sheet(self, writer: pd.ExcelWriter):
        """Escribe hoja de cambios detectados"""
        if not self.comparison:
            return
        
        # Cambios generales
        changes_df = self.comparison.get_changes_df()
        if len(changes_df) > 0:
            changes_df.to_excel(writer, sheet_name='Cambios', index=False)
        
        # Nuevas oportunidades
        if len(self.comparison.new_opportunities) > 0:
            self.comparison.new_opportunities.to_excel(
                writer, sheet_name='Nuevas', index=False
            )
        
        # Oportunidades eliminadas
        if len(self.comparison.removed_opportunities) > 0:
            self.comparison.removed_opportunities.to_excel(
                writer, sheet_name='Eliminadas', index=False
            )
    
    def _write_data_sheet(self, writer: pd.ExcelWriter):
        """Escribe hoja con todos los datos"""
        self.df.to_excel(writer, sheet_name='Datos_Completos', index=False)
    
    def _write_responsible_summary(self, writer: pd.ExcelWriter, 
                                    responsible: str, df: pd.DataFrame):
        """Escribe resumen para un responsable específico"""
        summary_data = [
            ['REPORTE DE OPORTUNIDADES', ''],
            ['Responsable', responsible],
            ['Fecha', datetime.now().strftime('%Y-%m-%d')],
            ['', ''],
            ['Total Oportunidades', len(df)],
            ['Valor Total USD', f"${df[COLUMNS['usd']].sum():,.2f}"],
            ['', ''],
            ['Por País:', ''],
        ]
        
        for market, count in df[COLUMNS['market']].value_counts().items():
            usd = df[df[COLUMNS['market']]==market][COLUMNS['usd']].sum()
            summary_data.append([f"  {market}", f"{count} opps (${usd:,.2f})"])
        
        pd.DataFrame(summary_data, columns=['Métrica', 'Valor']).to_excel(
            writer, sheet_name='Resumen', index=False
        )


def generate_weekly_report(df: pd.DataFrame, 
                          comparison: Optional[ComparisonResult] = None,
                          output_path: Optional[Path] = None) -> Path:
    """
    Función de conveniencia para generar reporte semanal
    
    Args:
        df: DataFrame con oportunidades
        comparison: Resultado de comparación (opcional)
        output_path: Ruta de salida (opcional)
        
    Returns:
        Path al archivo generado
    """
    generator = ExcelReportGenerator(df, comparison)
    return generator.generate_weekly_report(output_path)


def generate_responsible_reports(df: pd.DataFrame,
                                 output_dir: Optional[Path] = None) -> List[Path]:
    """
    Función de conveniencia para generar reportes por responsable
    
    Args:
        df: DataFrame con oportunidades
        output_dir: Directorio de salida
        
    Returns:
        Lista de paths a archivos generados
    """
    generator = ExcelReportGenerator(df)
    return generator.generate_responsible_reports(output_dir)
